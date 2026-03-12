"""
job_seek/exporters/csv_xlsx.py
--------------------------------
Export job records to CSV and XLSX.

Usage
-----
    from job_seek.exporters.csv_xlsx import export_all

    export_all(
        raw_records=all_jobs,
        filtered_records=kept_jobs,
        output_dir=Path("data"),
    )

Output files
------------
  data/raw_jobs.csv              — every scraped job (before filtering)
  data/filtered_cs_jobs.csv      — jobs with relevance_score >= threshold
  data/filtered_cs_jobs.xlsx     — same data, formatted as an Excel workbook
"""

from __future__ import annotations

import csv
from pathlib import Path
from typing import Any

from job_seek.utils.logger import get_logger
from job_seek.utils.normalizer import OUTPUT_COLUMNS

log = get_logger(__name__)

# ── File names ────────────────────────────────────────────────────────────────
RAW_CSV_NAME = "raw_jobs.csv"
FILTERED_CSV_NAME = "filtered_cs_jobs.csv"
FILTERED_XLSX_NAME = "filtered_cs_jobs.xlsx"

# ── XLSX styling constants ────────────────────────────────────────────────────
HEADER_FILL_COLOR = "1F4E79"   # dark blue
HEADER_FONT_COLOR = "FFFFFF"   # white
KEEP_FILL_COLOR = "E2EFDA"     # light green
REJECT_FILL_COLOR = "FCE4D6"   # light red


def _to_row(record: dict[str, Any]) -> list[str]:
    """Convert a record dict to a list of strings aligned with OUTPUT_COLUMNS."""
    return [str(record.get(col, "Unknown")) for col in OUTPUT_COLUMNS]


def write_csv(records: list[dict[str, Any]], path: Path) -> None:
    """Write records to a UTF-8 CSV file with OUTPUT_COLUMNS as header."""
    path.parent.mkdir(parents=True, exist_ok=True)
    with open(path, "w", newline="", encoding="utf-8") as fh:
        writer = csv.writer(fh, quoting=csv.QUOTE_MINIMAL)
        writer.writerow(OUTPUT_COLUMNS)
        for rec in records:
            writer.writerow(_to_row(rec))
    log.info("CSV written: %s  (%d rows)", path, len(records))


def write_xlsx(records: list[dict[str, Any]], path: Path) -> None:
    """
    Write records to a formatted Excel workbook.
    Requires openpyxl >= 3.1.
    """
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        from openpyxl.utils import get_column_letter
    except ImportError:
        log.error("openpyxl not installed — skipping XLSX export.  Run: pip install openpyxl")
        return

    path.parent.mkdir(parents=True, exist_ok=True)

    wb = Workbook()
    ws = wb.active
    ws.title = "Filtered CS Jobs"

    # ── Header row ────────────────────────────────────────────────────────────
    header_fill = PatternFill("solid", fgColor=HEADER_FILL_COLOR)
    header_font = Font(bold=True, color=HEADER_FONT_COLOR, size=10)
    thin = Side(style="thin", color="AAAAAA")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    for col_idx, col_name in enumerate(OUTPUT_COLUMNS, start=1):
        cell = ws.cell(row=1, column=col_idx, value=col_name)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = border

    # ── Data rows ─────────────────────────────────────────────────────────────
    keep_fill = PatternFill("solid", fgColor=KEEP_FILL_COLOR)
    reject_fill = PatternFill("solid", fgColor=REJECT_FILL_COLOR)
    data_font = Font(size=10)

    for row_idx, rec in enumerate(records, start=2):
        is_keep = str(rec.get("keep_or_reject", "Reject")).lower() == "keep"
        row_fill = keep_fill if is_keep else reject_fill

        for col_idx, col_name in enumerate(OUTPUT_COLUMNS, start=1):
            value = rec.get(col_name, "Unknown")

            # Convert numeric relevance_score to int
            if col_name == "relevance_score":
                try:
                    value = int(value)
                except (TypeError, ValueError):
                    value = 0

            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.fill = row_fill
            cell.font = data_font
            cell.alignment = Alignment(vertical="top", wrap_text=True)
            cell.border = border

    # ── Column widths ─────────────────────────────────────────────────────────
    col_widths = {
        "company_name": 20,
        "job_title": 35,
        "location": 20,
        "country": 12,
        "work_type": 12,
        "employment_type": 14,
        "seniority_level": 14,
        "department": 16,
        "job_url": 40,
        "source_site": 14,
        "date_posted": 14,
        "salary": 18,
        "visa_sponsorship": 16,
        "requires_degree": 14,
        "technical_category": 22,
        "relevance_score": 14,
        "keep_or_reject": 13,
        "rejection_reason": 25,
        "short_summary": 45,
        "required_skills": 40,
        "preferred_skills": 40,
    }

    for col_idx, col_name in enumerate(OUTPUT_COLUMNS, start=1):
        letter = get_column_letter(col_idx)
        ws.column_dimensions[letter].width = col_widths.get(col_name, 16)

    # ── Freeze header row + auto-filter ───────────────────────────────────────
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions

    # ── Summary sheet ─────────────────────────────────────────────────────────
    ws_summary = wb.create_sheet(title="Summary")
    total = len(records)
    kept_count = sum(1 for r in records if str(r.get("keep_or_reject", "")).lower() == "keep")
    rejected_count = total - kept_count

    # Count by technical_category
    from collections import Counter
    cat_counts = Counter(
        str(r.get("technical_category", "Unknown"))
        for r in records
        if str(r.get("keep_or_reject", "")).lower() == "keep"
    )

    summary_data = [
        ["Metric", "Value"],
        ["Total scraped (after dedup)", total],
        ["Kept (CS/IT roles)", kept_count],
        ["Rejected", rejected_count],
        ["", ""],
        ["Category Breakdown", "Count"],
    ] + [[cat, count] for cat, count in cat_counts.most_common()]

    for row in summary_data:
        ws_summary.append(row)

    ws_summary.column_dimensions["A"].width = 35
    ws_summary.column_dimensions["B"].width = 12

    wb.save(path)
    log.info("XLSX written: %s  (%d rows + summary sheet)", path, len(records))


def export_all(
    raw_records: list[dict[str, Any]],
    filtered_records: list[dict[str, Any]],
    output_dir: Path | str = "data",
) -> dict[str, Path]:
    """
    Write all three output files.

    Parameters
    ----------
    raw_records      : all scraped + normalised records (pre-filter)
    filtered_records : kept records only
    output_dir       : directory to write files into

    Returns
    -------
    dict mapping logical name → actual Path written
    """
    out = Path(output_dir)
    out.mkdir(parents=True, exist_ok=True)

    paths: dict[str, Path] = {}

    # 1. Raw CSV
    raw_path = out / RAW_CSV_NAME
    write_csv(raw_records, raw_path)
    paths["raw_csv"] = raw_path

    # 2. Filtered CSV
    filtered_csv_path = out / FILTERED_CSV_NAME
    write_csv(filtered_records, filtered_csv_path)
    paths["filtered_csv"] = filtered_csv_path

    # 3. Filtered XLSX
    filtered_xlsx_path = out / FILTERED_XLSX_NAME
    write_xlsx(filtered_records, filtered_xlsx_path)
    paths["filtered_xlsx"] = filtered_xlsx_path

    log.info(
        "Export summary — raw: %d rows | filtered: %d rows | output dir: %s",
        len(raw_records), len(filtered_records), out.resolve(),
    )
    return paths
