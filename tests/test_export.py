"""
tests/test_export.py
----------------------
Unit tests for CSV and XLSX export.

Run with:
  pytest tests/test_export.py -v
"""

import csv
import pytest
from pathlib import Path
from job_seek.exporters.csv_xlsx import write_csv, write_xlsx, export_all
from job_seek.utils.normalizer import OUTPUT_COLUMNS


# ─────────────────────────────────────────────────────────────────────────────
# Fixtures
# ─────────────────────────────────────────────────────────────────────────────

@pytest.fixture
def sample_records():
    return [
        {
            "company_name": "Stripe",
            "job_title": "Backend Engineer",
            "location": "Remote",
            "country": "USA",
            "work_type": "Remote",
            "employment_type": "Full-time",
            "seniority_level": "Senior",
            "department": "Engineering",
            "job_url": "https://jobs.lever.co/stripe/001",
            "source_site": "Lever",
            "date_posted": "2024-01-15",
            "salary": "$180k–$220k",
            "visa_sponsorship": "Yes",
            "requires_degree": "No",
            "technical_category": "Backend Engineering",
            "relevance_score": 85,
            "keep_or_reject": "Keep",
            "rejection_reason": "",
            "short_summary": "Build scalable payment infrastructure.",
            "required_skills": "Python; Go; PostgreSQL; AWS",
            "preferred_skills": "Kafka; Redis",
        },
        {
            "company_name": "Notion",
            "job_title": "Account Executive",
            "location": "New York, NY",
            "country": "USA",
            "work_type": "On-site",
            "employment_type": "Full-time",
            "seniority_level": "Mid",
            "department": "Sales",
            "job_url": "https://boards.greenhouse.io/notion/002",
            "source_site": "Greenhouse",
            "date_posted": "2024-01-16",
            "salary": "Unknown",
            "visa_sponsorship": "Unknown",
            "requires_degree": "Unknown",
            "technical_category": "Non-technical",
            "relevance_score": 10,
            "keep_or_reject": "Reject",
            "rejection_reason": "Sales role",
            "short_summary": "Close deals for Notion's enterprise tier.",
            "required_skills": "",
            "preferred_skills": "",
        },
    ]


@pytest.fixture
def output_dir(tmp_path):
    return tmp_path / "output"


# ─────────────────────────────────────────────────────────────────────────────
# CSV tests
# ─────────────────────────────────────────────────────────────────────────────

class TestWriteCsv:

    def test_creates_file(self, sample_records, output_dir):
        path = output_dir / "test.csv"
        write_csv(sample_records, path)
        assert path.exists()

    def test_header_matches_output_columns(self, sample_records, output_dir):
        path = output_dir / "test.csv"
        write_csv(sample_records, path)
        with open(path, newline="", encoding="utf-8") as f:
            reader = csv.reader(f)
            header = next(reader)
        assert header == OUTPUT_COLUMNS

    def test_row_count_correct(self, sample_records, output_dir):
        path = output_dir / "test.csv"
        write_csv(sample_records, path)
        with open(path, newline="", encoding="utf-8") as f:
            rows = list(csv.reader(f))
        assert len(rows) == len(sample_records) + 1  # +1 for header

    def test_empty_records_writes_header_only(self, output_dir):
        path = output_dir / "empty.csv"
        write_csv([], path)
        with open(path, newline="", encoding="utf-8") as f:
            rows = list(csv.reader(f))
        assert len(rows) == 1
        assert rows[0] == OUTPUT_COLUMNS

    def test_values_written_correctly(self, sample_records, output_dir):
        path = output_dir / "test.csv"
        write_csv(sample_records, path)
        with open(path, newline="", encoding="utf-8") as f:
            reader = csv.DictReader(f)
            rows = list(reader)
        assert rows[0]["company_name"] == "Stripe"
        assert rows[0]["job_title"] == "Backend Engineer"
        assert rows[1]["keep_or_reject"] == "Reject"

    def test_creates_parent_directories(self, sample_records, tmp_path):
        path = tmp_path / "nested" / "deep" / "test.csv"
        write_csv(sample_records, path)
        assert path.exists()

    def test_missing_fields_default_to_unknown(self, output_dir):
        minimal = [{"job_title": "Software Engineer", "company_name": "Acme"}]
        path = output_dir / "minimal.csv"
        write_csv(minimal, path)
        with open(path, newline="", encoding="utf-8") as f:
            reader = csv.DictReader(f)
            row = next(reader)
        assert row["location"] == "Unknown"
        assert row["salary"] == "Unknown"


# ─────────────────────────────────────────────────────────────────────────────
# XLSX tests
# ─────────────────────────────────────────────────────────────────────────────

class TestWriteXlsx:

    def test_creates_file(self, sample_records, output_dir):
        path = output_dir / "test.xlsx"
        write_xlsx(sample_records, path)
        assert path.exists()

    def test_file_is_valid_xlsx(self, sample_records, output_dir):
        pytest.importorskip("openpyxl")
        from openpyxl import load_workbook
        path = output_dir / "test.xlsx"
        write_xlsx(sample_records, path)
        wb = load_workbook(path)
        assert "Filtered CS Jobs" in wb.sheetnames

    def test_header_row_in_xlsx(self, sample_records, output_dir):
        pytest.importorskip("openpyxl")
        from openpyxl import load_workbook
        path = output_dir / "test.xlsx"
        write_xlsx(sample_records, path)
        wb = load_workbook(path)
        ws = wb["Filtered CS Jobs"]
        header = [ws.cell(row=1, column=i).value for i in range(1, len(OUTPUT_COLUMNS) + 1)]
        assert header == OUTPUT_COLUMNS

    def test_data_row_count(self, sample_records, output_dir):
        pytest.importorskip("openpyxl")
        from openpyxl import load_workbook
        path = output_dir / "test.xlsx"
        write_xlsx(sample_records, path)
        wb = load_workbook(path)
        ws = wb["Filtered CS Jobs"]
        assert ws.max_row == len(sample_records) + 1  # +1 header

    def test_summary_sheet_exists(self, sample_records, output_dir):
        pytest.importorskip("openpyxl")
        from openpyxl import load_workbook
        path = output_dir / "test.xlsx"
        write_xlsx(sample_records, path)
        wb = load_workbook(path)
        assert "Summary" in wb.sheetnames


# ─────────────────────────────────────────────────────────────────────────────
# export_all tests
# ─────────────────────────────────────────────────────────────────────────────

class TestExportAll:

    def test_creates_all_three_files(self, sample_records, output_dir):
        kept = [r for r in sample_records if r["keep_or_reject"] == "Keep"]
        paths = export_all(
            raw_records=sample_records,
            filtered_records=kept,
            output_dir=output_dir,
        )
        assert paths["raw_csv"].exists()
        assert paths["filtered_csv"].exists()
        assert paths["filtered_xlsx"].exists()

    def test_raw_csv_has_all_records(self, sample_records, output_dir):
        kept = [r for r in sample_records if r["keep_or_reject"] == "Keep"]
        paths = export_all(sample_records, kept, output_dir)
        with open(paths["raw_csv"], newline="", encoding="utf-8") as f:
            rows = list(csv.reader(f))
        assert len(rows) == len(sample_records) + 1

    def test_filtered_csv_has_only_kept_records(self, sample_records, output_dir):
        kept = [r for r in sample_records if r["keep_or_reject"] == "Keep"]
        paths = export_all(sample_records, kept, output_dir)
        with open(paths["filtered_csv"], newline="", encoding="utf-8") as f:
            rows = list(csv.reader(f))
        assert len(rows) == len(kept) + 1

    def test_returns_correct_path_keys(self, sample_records, output_dir):
        kept = [r for r in sample_records if r["keep_or_reject"] == "Keep"]
        paths = export_all(sample_records, kept, output_dir)
        assert set(paths.keys()) == {"raw_csv", "filtered_csv", "filtered_xlsx"}
